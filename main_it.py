import os
import torch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer, AutoModel, AutoModelForCausalLM
from sentence_transformers import SentenceTransformer
import fasttext

from utils_it import load_main_csv, load_psyn_xlsx, load_prand_xlsx
from for_probes_it import sim_scores, has_nc

MAIN_CSV   = "MWEs_completo_tagged.csv"

PSYN_XLSX  = "frasi_PSyn_corretto.xlsx"

NAT_PRAND_PATHS = [
    "frasi_naturali_prand_1.xlsx",
    "frasi_naturali_prand_2.xlsx",
    "frasi_naturali_prand_3.xlsx",
    "frasi_naturali_prand_4.xlsx",
    "frasi_naturali_prand_5.xlsx",
]
NEUT_PRAND_PATHS = [
    "frasi_neutre_prand_1.xlsx",
    "frasi_neutre_prand_2.xlsx",
    "frasi_neutre_prand_3.xlsx",
    "frasi_neutre_prand_4.xlsx",
    "frasi_neutre_prand_5.xlsx",
]

FASTTEXT_MODEL_PATH = "cc.it.300.bin"

OUTPUT_DIR = "results"

MODEL_REGISTRY = {
    "BERTita": {
        "family":   "bert",
        "model_id": "dbmdz/bert-base-italian-xxl-cased",
    },
    "mBERT": {
        "family":   "bert",
        "model_id": "google-bert/bert-base-multilingual-cased",
    },
    "mDistilBERT": {
        "family":   "bert",
        "model_id": "distilbert/distilbert-base-multilingual-cased",
    },
    "mSBERT": {
        "family":   "sbert",
        "model_id": "distiluse-base-multilingual-cased-v1",
    },
    "fastText": {
        "family":   "fasttext",
        "model_id": FASTTEXT_MODEL_PATH,
    },
    "Minerva": {
        "family":   "minerva",
        "model_id": "sapienzanlp/Minerva-3B-base-v1.0",
    },
}

PROBES = ["PSyn", "PComp", "PWordsSyn", "PRand"]

if torch.backends.mps.is_available():
    device = "mps"
elif torch.cuda.is_available():
    device = "cuda"
else:
    device = "cpu"
print(f"Device: {device}")

def write_xlsx(risultati, output_path, sheet_title):
    df_out = pd.DataFrame(risultati)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  

    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alt_fill    = PatternFill("solid", start_color="EEF4FB", end_color="EEF4FB")
    white_fill  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    data_font   = Font(name="Arial", size=9)

    cols = list(df_out.columns)
    for c, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, record in enumerate(risultati, start=2):
        bg = alt_fill if r % 2 == 0 else white_fill
        for c, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c, value=record.get(col_name, ""))
            cell.font      = data_font
            cell.fill      = bg
            cell.alignment = Alignment(
                horizontal="center" if c > 1 else "left",
                vertical="center"
            )

    ws.column_dimensions["A"].width = 28
    for c in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "B2"

    wb.save(output_path)
    print(f"  Saved: {output_path}")

def load_model(model_name, config):

    family   = config["family"]
    model_id = config["model_id"]

    print(f"\nLoading {model_name} ({family}) ...")

    if family == "bert":
        tokenizer = AutoTokenizer.from_pretrained(model_id)
        model     = AutoModel.from_pretrained(model_id, output_hidden_states=True).to(device)
        model.eval()
        return model, tokenizer

    elif family == "minerva":
        tokenizer = AutoTokenizer.from_pretrained(model_id)
        model     = AutoModelForCausalLM.from_pretrained(
            model_id, torch_dtype=torch.float16, output_hidden_states=True
        ).to(device)
        model.eval()
        return model, tokenizer

    elif family == "sbert":
        model = SentenceTransformer(model_id)
        return model, None

    elif family == "fasttext":
        model = fasttext.load_model(model_id)
        return model, None

    else:
        raise ValueError(f"Unknown model family: {family}")

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Loading data...")
    df_main  = load_main_csv(MAIN_CSV)
    df_psyn  = load_psyn_xlsx(PSYN_XLSX)
    nat_dfs, neut_dfs = load_prand_xlsx(NAT_PRAND_PATHS, NEUT_PRAND_PATHS)
    print(f"  {len(df_main)} MWEs loaded from {MAIN_CSV}")

    for model_name, config in MODEL_REGISTRY.items():

        model, tokenizer = load_model(model_name, config)

        for probe in PROBES:

            print(f"\n[{model_name}] probe={probe}")

            risultati = sim_scores(
                probe      = probe,
                df_main    = df_main,
                model_name = model_name,
                model      = model,
                tokenizer  = tokenizer,
                device     = device,
                df_psyn    = df_psyn   if probe == "PSyn"  else None,
                nat_dfs    = nat_dfs   if probe == "PRand" else None,
                neut_dfs   = neut_dfs  if probe == "PRand" else None,
            )

            output_path  = os.path.join(OUTPUT_DIR, f"Sim_{probe}_{model_name}.xlsx")
            sheet_title  = f"Sim({probe}) {model_name}"
            write_xlsx(risultati, output_path, sheet_title)

        del model
        if device in ("cuda", "mps"):
            torch.cuda.empty_cache() if device == "cuda" else torch.mps.empty_cache()


if __name__ == "__main__":
    main()
